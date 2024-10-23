import openpyxl
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from flask import Flask, render_template, request, flash, redirect, url_for, jsonify
import os
from werkzeug.utils import secure_filename
import google.generativeai as genai

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'  # Replace with a real secret key
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Configure the Gemini API
genai.configure(api_key='AIzaSyAYaBIKu3m-LcHGj-11tBJpmo6yMKU-NB4')  # Replace with your actual API key
model = genai.GenerativeModel('gemini-1.5-pro')

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def send_mass_email(excel_file, email_subject, email_content, sender_email, sender_password):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    server = smtplib.SMTP(smtp_server, smtp_port)
    server.starttls()
    server.login(sender_email, sender_password)

    for row in sheet.iter_rows(min_row=1, values_only=True):
        recipient_email = row[0]

        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = recipient_email
        message["Subject"] = email_subject

        message.attach(MIMEText(email_content, "plain"))
        server.send_message(message)
        print(f"Email sent to: {recipient_email}")

    server.quit()
    print("Mass email sending completed.")

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            email_subject = request.form['email_subject']
            email_content = request.form['email_content']
            sender_email = request.form['sender_email']
            sender_password = request.form['sender_password']
            
            try:
                send_mass_email(filepath, email_subject, email_content, sender_email, sender_password)
                flash('Mass email sent successfully!')
            except Exception as e:
                flash(f'An error occurred: {str(e)}')
            
            os.remove(filepath)  # Remove the uploaded file after processing
            return redirect(url_for('index'))
    return render_template('index.html')

@app.route('/optimize_content', methods=['POST'])
def optimize_content():
    content = request.json['content']
    prompt = f"""
    Enhance the given text. If its betwee 60 to 150 words already, add some words and bits or paraphrase wherever necessary with the optimized content , without much affecting the texts originality. 
    If The text is smaller than that optimise whatever necessary. If brand names or somethin is not specified, add some dummy names in context to the 
    paragraph. 

    Dont add ## or * or ** to make things bold or title like. Dont keep any title.
    {content}
    
    Please ensure the optimized content:
    1. Is SEO-friendly
    2. Follows sales funnel principles
    3. Uses catchy words to grab attention
    4. Is at least 150 words long (add relevant content if needed)
    5. Maintains the original message's intent
    """
    
    try:
        response = model.generate_content(prompt)
        optimized_content = response.text
        return jsonify({'optimized_content': optimized_content})
    except Exception as e:
        print(f"Error in Gemini API call: {str(e)}")
        return jsonify({'error': 'Failed to optimize content'}), 500

@app.route('/send_mass_email', methods=['POST'])
def send_mass_email_route():  # Renamed this function
    try:
        file = request.files['file']
        email_subject = request.form['email_subject']
        email_content = request.form['email_content']
        sender_email = request.form['sender_email']
        sender_password = request.form['sender_password']

        # Save the uploaded file
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)

        # Process the Excel file and send emails
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            recipient_email = row[0]  # Assuming email is in the first column
            
            # Create the email
            msg = MIMEMultipart()
            msg['From'] = sender_email
            msg['To'] = recipient_email
            msg['Subject'] = email_subject
            msg.attach(MIMEText(email_content, 'plain'))

            # Send the email
            with smtplib.SMTP('smtp.gmail.com', 587) as server:
                server.starttls()
                server.login(sender_email, sender_password)
                server.send_message(msg)

        # Clean up: remove the uploaded file
        os.remove(file_path)

        return jsonify({'message': 'Mass email sent successfully!'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    app.run(debug=True)
